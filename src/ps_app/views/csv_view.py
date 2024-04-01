import os
from PyQt5 import QtWidgets
import openpyxl
import pickle
import base64


class ClusterData(QtWidgets.QTableWidget):
    def __init__(self, cluster_map, msa, low_entropy, column_map):
        super().__init__()
        self.setRowCount(0)
        self.setColumnCount(3)
        self.cluster_map = cluster_map
        self.msa = msa
        self.low_entropy = low_entropy
        self.column_map = column_map

        file = [[str(k), str(v[0]), str(v[1])] for k, v in self.cluster_map.items()]
        file.insert(0, ['Cluster', 'SR Mode', 'Discovered'])

        for row_data in file:
            row = self.rowCount()
            self.insertRow(row)
            for col, each in enumerate(row_data):
                entry = row_data[col]
                item = QtWidgets.QTableWidgetItem(entry)
                self.setItem(row, col, item)

    def save_sheet(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Excel", os.getenv('HOME'), 'Excel (*.xlsx)')
        if path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            # Write results to sheet
            for row in range(self.rowCount()):
                row_data = []
                for column in range(self.columnCount()):
                    item = self.item(row, column)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append('')
                ws.append(row_data)

            if self.msa is not None:
                # Entropy Sheet
                if self.low_entropy is not None:
                    low_entropy_res = self.msa[self.low_entropy]
                    entropy_sheet = wb.create_sheet(title="Low Entropy")
                    entropy_sheet.append(low_entropy_res.columns.to_list())
                    for index, row in low_entropy_res.iterrows():
                        row_data = row.tolist()
                        entropy_sheet.append(row_data)

                # Cluster sheets
                cluster_lengths = sorted(set([len(k) for k, _ in self.cluster_map.items()]))
                for n in cluster_lengths:
                    if n > 10:
                        continue

                    title = ""
                    if n == 2:
                        title = "Pairwise"
                    else:
                        title = ordinal(n) + " order"

                    sheet = wb.create_sheet(title=title)
                    n_tuples = [k for k, v in self.cluster_map.items() if len(k) == n]
                    for idx, p in enumerate(n_tuples):
                        # Calculate the starting column for the new tuple
                        start_column = idx * (n+1) + 1

                        # Insert each pandas Series as a column next to each other
                        for i in range(n):
                            sheet.cell(row=1, column=start_column+i, value=p[i])
                        sheet.cell(row=1, column=start_column + n, value='')

                        # Insert values
                        df_list = [self.msa[column] for column in p]
                        max_len = max([len(df) for df in df_list])
                        for row_idx in range(max_len):
                            for i, df in enumerate(df_list):
                                value = df.iloc[row_idx] if row_idx < len(df) else ''
                                sheet.cell(row=row_idx+2, column=start_column + i, value=value)
                            sheet.cell(row=row_idx+2, column=start_column + len(p), value='')

            serialized = pickle.dumps({'cluster_map': self.cluster_map, 'msa': self.msa, 'low_entropy':
                                       self.low_entropy, 'column_map': self.column_map})
            encoded = base64.b64encode(serialized).decode('utf-8')

            chunks = [encoded[i:i+255] for i in range(0, len(encoded), 255)]
            for i, chunk in enumerate(chunks):
                wb.custom_doc_props.append(openpyxl.packaging.custom.StringProperty(name=f"pcviewer#{i}", value=chunk))
            wb.save(path)


def ordinal(n):
    if 10 <= n <= 20:
        return str(n) + 'th'
    else:
        last = n % 10
        if last == 1:
            return str(n) + 'st'
        elif last == 2:
            return str(n) + 'nd'
        elif last == 3:
            return str(n) + 'rd'
        else:
            return str(n) + 'th'
