import os
import csv
from PyQt5 import QtWidgets
import openpyxl

class ClusterData(QtWidgets.QTableWidget):
    def __init__(self, path, df, low_entropy):
        super().__init__()
        self.setRowCount(0)
        self.setColumnCount(3)
        self.cluster_map = path
        self.df = df
        self.low_entropy = low_entropy
        if type(self.cluster_map) is dict:
            file = [[str(k), str(v[0]), str(v[1])] for k, v in self.cluster_map.items()]
            file.insert(0, ['Cluster', 'SR Mode', 'Discovered'])
        else:
            data = open(self.cluster_map)
            file = csv.reader(data)
            print(file)

        for row_data in file:
            row = self.rowCount()
            self.insertRow(row)
            for col, each in enumerate(row_data):
                entry = row_data[col]
                item = QtWidgets.QTableWidgetItem(entry)
                self.setItem(row, col, item)

    def save_sheet(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Excel", os.getenv('HOME'), 'Excel (*.xlsx)')
        if path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            # Write results to sheet
            for row in range(self.rowCount()):
                row_data = []
                for column in range(self.columnCount()):
                    item = self.item(row, column)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append('')
                ws.append(row_data)

            if self.df is not None:
                # Pairwise Sheet
                pairwise_sheet = wb.create_sheet(title="Pairwise")
                pairs = [k for k, v in self.cluster_map.items() if len(k) == 2]
                for idx, p in enumerate(pairs):
                    column_1, column_2 = p
                    df1, df2 = self.df[column_1], self.df[column_2]
                    start_column = idx * 3 + 1  # Calculate the starting column for the new pair

                    # Insert each pandas Series as a column next to each other
                    pairwise_sheet.cell(row=1, column=start_column, value=column_1)
                    pairwise_sheet.cell(row=1, column=start_column + 1, value=column_2)
                    pairwise_sheet.cell(row=1, column=start_column + 2, value='')

                    for row_idx in range(max(len(df1), len(df2))):
                        value_1 = df1.iloc[row_idx] if row_idx < len(df1) else ''
                        value_2 = df2.iloc[row_idx] if row_idx < len(df2) else ''
                        pairwise_sheet.cell(row=row_idx + 2, column=start_column, value=value_1)
                        pairwise_sheet.cell(row=row_idx + 2, column=start_column + 1, value=value_2)
                        pairwise_sheet.cell(row=row_idx + 2, column=start_column + 2, value='')

                # Entropy Sheet
                if self.low_entropy is not None:
                    low_entropy_res = self.df[self.low_entropy]
                    entropy_sheet = wb.create_sheet(title="Low Entropy")
                    entropy_sheet.append(low_entropy_res.columns.to_list())
                    for index, row in low_entropy_res.iterrows():
                        row_data = row.tolist()
                        entropy_sheet.append(row_data)

            wb.save(path)
